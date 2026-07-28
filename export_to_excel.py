#!/usr/bin/env python3
"""
SRMS Export - Fixed version with proper data handling
"""

import json
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def flatten_value(value):
    """Convert complex values to Excel-friendly format"""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        # Convert dict to JSON string
        return json.dumps(value)
    if isinstance(value, list):
        # Convert list to comma-separated string
        return ", ".join(str(item) for item in value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    # For any other type, convert to string
    return str(value)

def export_to_excel():
    """Main export function"""
    try:
        print("📡 Fetching data from API...")
        response = requests.get('https://srms-api.jamesmoriarty.in/api/google-sheets/export-json')
        
        if response.status_code != 200:
            print(f"❌ API error: {response.status_code}")
            return False
        
        data = response.json()
        if data.get('status') != 'success':
            print(f"❌ API returned error: {data.get('message', 'Unknown error')}")
            return False
        
        tables = data.get('data', {})
        print(f"✅ Retrieved {len(tables)} tables: {', '.join(tables.keys())}")
        
        # Create workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        for table_name, rows in tables.items():
            if not rows:
                print(f"⚠️  Table '{table_name}' is empty, skipping")
                continue
            
            print(f"📊 Processing table: {table_name} ({len(rows)} rows)")
            
            # Create sheet with valid name
            sheet_name = table_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Get headers
            headers = list(rows[0].keys())
            
            # Write headers with formatting
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Write data rows with flattening
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header)
                    flat_value = flatten_value(value)
                    ws.cell(row=row_idx, column=col_idx, value=flat_value)
            
            # Auto-size columns
            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row_idx in range(2, min(len(rows) + 2, 100)):  # Check first 100 rows
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_len = max(max_len, len(str(cell_value)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
            
            print(f"✅ Table '{table_name}' exported")
        
        # Save file
        filename = '/tmp/srms_export.xlsx'
        wb.save(filename)
        print(f"✅ Excel file saved: {filename}")
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = export_to_excel()
    sys.exit(0 if success else 1)
