#!/usr/bin/env python3
import json
import sys
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import os

# Suppress all warnings and non-essential output
import warnings
warnings.filterwarnings("ignore")

def main():
    try:
        # Fetch data
        response = requests.get('https://srms-api.jamesmoriarty.in/api/google-sheets/export-json', timeout=30)
        if response.status_code != 200:
            sys.exit(1)
        
        data = response.json()
        if data.get('status') != 'success':
            sys.exit(1)
        
        tables = data.get('data', {})
        
        # Create workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        for table_name, rows in tables.items():
            ws = wb.create_sheet(title=table_name[:31])
            if rows and len(rows) > 0:
                headers = list(rows[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row.get(header, '')
                        if value is None:
                            value = ''
                        ws.cell(row=row_idx, column=col_idx, value=value)
                for col in ws.columns:
                    max_len = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        
        # Save to base64
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        base64_data = base64.b64encode(output.getvalue()).decode('utf-8')
        
        # Print ONLY the base64 data (no other output)
        sys.stdout.write(base64_data)
        sys.stdout.flush()
        
    except Exception as e:
        # On error, print nothing (or an empty string) to avoid corruption
        sys.exit(1)

if __name__ == "__main__":
    main()
