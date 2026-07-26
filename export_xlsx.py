#!/usr/bin/env python3
import json
import sys
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import requests

# Fetch data
response = requests.get('https://srms-api.jamesmoriarty.in/api/google-sheets/export-json')
if response.status_code != 200:
    sys.exit(1)

data = response.json()
if data.get('status') != 'success':
    sys.exit(1)

tables = data.get('data', {})

# Create workbook
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

for table_name, rows in tables.items():
    ws = wb.create_sheet(title=table_name[:31])
    if rows and len(rows) > 0:
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header, '')
                if value is None:
                    value = ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

# Save to base64 and print (this is the ONLY output)
output = io.BytesIO()
wb.save(output)
output.seek(0)
base64_data = base64.b64encode(output.getvalue()).decode('utf-8')
print(base64_data)
