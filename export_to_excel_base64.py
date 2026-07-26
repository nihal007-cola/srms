#!/usr/bin/env python3
import json
import sys
import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import requests

def create_excel_base64():
    """Create Excel file in memory and return as base64 encoded string"""
    
    # Fetch data from the API
    response = requests.get('https://srms-api.jamesmoriarty.in/api/google-sheets/export-json')
    
    if response.status_code != 200:
        print(f"❌ Error fetching data: {response.status_code}", file=sys.stderr)
        return None
    
    data = response.json()
    
    if data.get('status') != 'success':
        print(f"❌ API error: {data.get('message', 'Unknown error')}", file=sys.stderr)
        return None
    
    tables = data.get('data', {})
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for table_name, rows in tables.items():
        # Create a new sheet
        ws = wb.create_sheet(title=table_name[:31])
        
        if rows and len(rows) > 0:
            # Get column headers
            headers = list(rows[0].keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header, '')
                    if value is None:
                        value = ''
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    # Save to memory (BytesIO)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Encode as base64
    base64_data = base64.b64encode(output.getvalue()).decode('utf-8')
    return base64_data

if __name__ == "__main__":
    result = create_excel_base64()
    if result:
        print(result)
    else:
        sys.exit(1)
